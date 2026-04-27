"""
학술대회 프로그램북 생성 - 메인 실행 스크립트

사용법:
    python generate_program.py --init          # 최초 1회: input_raw.xlsx → 작업 파일 생성
    python generate_program.py --step 1        # Step 1 실행 (데이터 정제)
    python generate_program.py --step 2        # Step 2 실행 (랩 매핑)
    python generate_program.py --step 3        # Step 3 실행 (제약조건 + 좌장/Invited 템플릿)
    python generate_program.py --step 4        # Step 4 실행 (배정 도우미 + 검증)
    python generate_program.py --step 6        # Step 6 실행 (최종 시트 생성)
    python generate_program.py --step all      # Step 1, 2, 3, 4, 6 순차 실행

작업 파일: program_book_working.xlsx (모든 시트가 한 파일에)

사용자 편집 시트 (재실행해도 보존됨):
  - 시트2             : 원본 submission 데이터
  - 입력_좌장Invited  : Step 3가 만든 템플릿. Invited 마킹용
  - 설정_시간대       : Step 4가 만든 템플릿. 6개 시간 슬롯
  - 입력_논문배정     : Step 4가 만든 템플릿. 논문별 날짜/시간/발표장 배정
  - 입력_세션좌장     : Step 6가 만든 템플릿. 세션명/좌장 입력

자동 생성 시트 (매번 덮어씀):
  - _step1_정제데이터, _step1_예외항목
  - _step2_랩매핑, _step2_소속정규화, _step2_랩요약
  - _step3_플래그
  - _step4_배정현황, _step4_충돌경고, _step4_소속축약
  - 5월7일(목), 5월8일(금), 발표 세부 일정, 구두 발표 명단, 포스터 발표 명단
"""

import argparse
from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from scripts import (
    step1_cleaning, step2_lab_mapping, step3_flagging,
    step4_assignment, step6_formatting,
)


WORKING_FILE = 'program_book_working.xlsx'
RAW_INPUT_FILE = 'input_raw.xlsx'
RAW_SHEET = '시트2'


# ---------------------------------------------------------------------------
# 파일 입출력 헬퍼
# ---------------------------------------------------------------------------

def load_raw_data():
    """시트2에서 원본 데이터 로드."""
    if not Path(WORKING_FILE).exists():
        raise FileNotFoundError(
            f"{WORKING_FILE}이 없습니다. --init 옵션으로 먼저 생성하세요."
        )
    return pd.read_excel(WORKING_FILE, sheet_name=RAW_SHEET, dtype=str)


def write_sheet(df, sheet_name, header=True):
    """작업 파일에 시트 쓰기 (있으면 덮어씀)."""
    mode = 'a' if Path(WORKING_FILE).exists() else 'w'
    kwargs = {'engine': 'openpyxl', 'mode': mode}
    if mode == 'a':
        kwargs['if_sheet_exists'] = 'replace'
    with pd.ExcelWriter(WORKING_FILE, **kwargs) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=header)


def apply_assignment_dropdowns(df_timeslots):
    """입력_논문배정 시트의 날짜/세션/발표장/슬롯순서/슬롯길이 컬럼에 드롭다운 적용."""
    wb = load_workbook(WORKING_FILE)
    if '입력_논문배정' not in wb.sheetnames:
        wb.close()
        return
    ws = wb['입력_논문배정']
    max_row = ws.max_row
    if max_row < 2:
        wb.close()
        return

    # 헤더 → 컬럼 번호
    col_idx = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            col_idx[str(h)] = c

    # 설정_시간대에서 세션 리스트 추출 (등장 순서 유지, 빈값 제외)
    sessions = []
    seen = set()
    if '세션' in df_timeslots.columns:
        for raw in df_timeslots['세션']:
            v = step4_assignment._clean_cell(raw)
            if v and v not in seen:
                sessions.append(v)
                seen.add(v)

    def add_list_dv(col_name, values, error_msg=None, enforce=True):
        if col_name not in col_idx or not values:
            return
        c = col_idx[col_name]
        letter = get_column_letter(c)
        formula = '"' + ','.join(str(v) for v in values) + '"'
        dv = DataValidation(
            type='list', formula1=formula, allow_blank=True,
            showErrorMessage=enforce,
        )
        if enforce:
            dv.errorTitle = '잘못된 값'
            dv.error = error_msg or f'{col_name}은(는) [{", ".join(str(v) for v in values)}] 중에서 선택'
        ws.add_data_validation(dv)
        dv.add(f'{letter}2:{letter}{max_row}')

    add_list_dv('날짜', ['목', '금'])
    add_list_dv('세션', sessions)
    add_list_dv('발표장', ['1', '2', '3', '4', '5', '6', '7'])
    # 슬롯순서: 구두는 1~5만 유효하나 포스터는 세션 내 순번(>5 가능) → 드롭다운은 유지하되 강제하지 않음
    add_list_dv('슬롯순서', ['1', '2', '3', '4', '5'], enforce=False)
    add_list_dv('슬롯길이', ['1', '2', '3', '4', '5'])

    wb.save(WORKING_FILE)
    wb.close()


def _display_width(text):
    """CJK 문자는 2칸, ASCII는 1칸으로 카운트."""
    if text is None:
        return 0
    w = 0
    for line in str(text).split('\n'):
        line_w = 0
        for c in line:
            line_w += 2 if ord(c) > 127 else 1
        if line_w > w:
            w = line_w
    return w


def autofit_columns(sheet_names=None, min_width=6, max_width=80):
    """지정된 시트들의 컬럼 폭을 내용에 맞춰 자동 조정."""
    if not Path(WORKING_FILE).exists():
        return
    wb = load_workbook(WORKING_FILE)
    targets = sheet_names or wb.sheetnames
    for sn in targets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_w = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                w = _display_width(cell.value)
                if w > max_w:
                    max_w = w
            if max_w == 0:
                continue
            width = max(min_width, min(max_width, max_w + 2))
            ws.column_dimensions[letter].width = width
    wb.save(WORKING_FILE)
    wb.close()


def snapshot_sheet_cells(sheet_name):
    """시트의 모든 비어있지 않은 셀을 {(row, col): value}로 스냅샷."""
    wb = load_workbook(WORKING_FILE)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]
    snap = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ''):
                snap[(r, c)] = v
    wb.close()
    return snap


def restore_empty_cells(sheet_name, snapshot):
    """새 시트가 비워둔 셀 위치에 한해, 스냅샷의 값을 복원.
    자동 생성된 셀(새 시트가 값을 쓴 위치)은 건드리지 않음 →
    사용자가 수작업으로 추가한 '빈 자리' 콘텐츠만 보존되는 효과.
    """
    if not snapshot:
        return 0
    wb = load_workbook(WORKING_FILE)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[sheet_name]
    restored = 0
    for (r, c), v in snapshot.items():
        if r > ws.max_row or c > ws.max_column:
            # 시트가 작아졌으면 확장
            pass
        new_v = ws.cell(row=r, column=c).value
        if new_v in (None, ''):
            ws.cell(row=r, column=c).value = v
            restored += 1
    wb.save(WORKING_FILE)
    return restored


def apply_merges(sheet_name, merges):
    """openpyxl로 지정 시트에 셀 병합 적용. merges: [(r1,c1,r2,c2), ...] 1-based."""
    if not merges:
        return
    wb = load_workbook(WORKING_FILE)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]
    for r1, c1, r2, c2 in merges:
        # 단일 셀이면 스킵
        if r1 == r2 and c1 == c2:
            continue
        try:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except Exception as e:
            print(f"  [주의] 병합 실패 ({sheet_name} {r1},{c1}-{r2},{c2}): {e}")
    wb.save(WORKING_FILE)
    wb.close()


def style_poster_session_sheet(sheet_name='포스터 세션 배치'):
    """포스터 세션 배치 시트에 기본 색/정렬/폭 적용."""
    if not Path(WORKING_FILE).exists():
        return

    field_fills = {
        '구조설계 및 CAE': 'D9D6FF',
        '인공지능/머신러닝': 'DDF8D2',
        '전산역학': 'FFF3BF',
        '멀티스케일': 'F2F2F2',
        '멀티피직스': 'FFF7CC',
        '통계기반': 'F5F5F5',
        '기타': 'DDF8D2',
    }
    default_fill = 'F7F7F7'
    title_fill = PatternFill(fill_type='solid', fgColor='F6C28B')
    thin = Side(style='thin', color='444444')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = load_workbook(WORKING_FILE)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    for row_idx in range(1, ws.max_row + 1):
        values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value not in (None, ''):
                values.append(str(cell.value).strip())

        first_text = values[0] if values else ''
        if first_text.startswith('포스터 세션'):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = title_fill
                cell.font = Font(bold=True)
                cell.alignment = center
                cell.border = border
            continue

        is_field_header = bool(values) and all('\n' not in value for value in values)
        if is_field_header and row_idx < ws.max_row:
            next_values = [
                ws.cell(row=row_idx + 1, column=col_idx).value
                for col_idx in range(1, ws.max_column + 1)
            ]
            next_has_poster = any(v not in (None, '') for v in next_values)
            if next_has_poster:
                current_fill = default_fill
                for col_idx in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=row_idx, column=col_idx)
                    header_text = '' if header_cell.value is None else str(header_cell.value).strip()
                    if header_text:
                        current_fill = field_fills.get(header_text, default_fill)
                    fill = PatternFill(fill_type='solid', fgColor=current_fill)
                    header_cell.fill = fill
                    header_cell.font = Font(bold=True)
                    header_cell.alignment = center
                    header_cell.border = border

                    poster_cell = ws.cell(row=row_idx + 1, column=col_idx)
                    poster_cell.fill = fill
                    poster_cell.alignment = center
                    poster_cell.border = border

                ws.row_dimensions[row_idx].height = 18
                ws.row_dimensions[row_idx + 1].height = 54
                continue

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value not in (None, ''):
                cell.alignment = center
                cell.border = border

    wb.save(WORKING_FILE)
    wb.close()


def sheet_exists(sheet_name):
    """작업 파일에 지정 시트가 존재하는지 확인."""
    if not Path(WORKING_FILE).exists():
        return False
    wb = load_workbook(WORKING_FILE, read_only=True)
    return sheet_name in wb.sheetnames


def init_working_file():
    """최초 실행: input_raw.xlsx → program_book_working.xlsx (시트2)."""
    if Path(WORKING_FILE).exists():
        print(f"[주의] {WORKING_FILE}이 이미 존재합니다. --init 건너뜀.")
        return
    if not Path(RAW_INPUT_FILE).exists():
        raise FileNotFoundError(f"{RAW_INPUT_FILE}이 없습니다.")
    raw = pd.read_excel(RAW_INPUT_FILE, sheet_name=0, dtype=str)
    with pd.ExcelWriter(WORKING_FILE, engine='openpyxl', mode='w') as writer:
        raw.to_excel(writer, sheet_name=RAW_SHEET, index=False)
    print(f"[완료] {WORKING_FILE} 생성 완료 ({RAW_SHEET} 시트에 {len(raw)}건 저장)")


# ---------------------------------------------------------------------------
# Step 실행
# ---------------------------------------------------------------------------

def run_step1():
    print("=" * 50)
    print("Step 1: 데이터 정제")
    print("=" * 50)
    raw = load_raw_data()
    print(f"  입력 행 수           : {len(raw)}")
    cleaned, exceptions, stats = step1_cleaning.clean(raw)
    print(f"  제목 조각 병합       : {stats['titles_merged']}건")
    print(f"  빈 행 제거           : {stats['empty_rows_removed']}건")
    print(f"  중복 제출 제거       : {stats['duplicates_removed']}건")
    print(f"  미등록/미납 제외     : {stats['unregistered_removed']}건")
    print(f"  정제 후 행 수        : {stats['rows_after_cleaning']}건")
    print(f"  예외항목 기록        : {stats['exceptions_logged']}건")
    write_sheet(cleaned, '_step1_정제데이터')
    write_sheet(exceptions, '_step1_예외항목')
    print(f"[완료] _step1_정제데이터, _step1_예외항목 시트 갱신 완료")


def run_step2():
    print("=" * 50)
    print("Step 2: 랩/그룹 식별")
    print("=" * 50)
    cleaned = pd.read_excel(WORKING_FILE, sheet_name='_step1_정제데이터', dtype=str)
    lab_df, norm_map_df, summary, stats = step2_lab_mapping.map_labs(cleaned)
    print(f"  정제 데이터 행 수       : {stats['n_papers']}")
    print(f"  고유 lab_id 수          : {stats['n_labs']}")
    print(f"  원본 소속 표기 수       : {stats['n_unique_aff_raw']}")
    print(f"  정규화 후 고유 소속 수  : {stats['n_unique_aff_norm']}")
    write_sheet(lab_df, '_step2_랩매핑')
    write_sheet(norm_map_df, '_step2_소속정규화')
    write_sheet(summary, '_step2_랩요약')
    print(f"[완료] _step2_랩매핑, _step2_소속정규화, _step2_랩요약 시트 갱신 완료")


def run_step3():
    print("=" * 50)
    print("Step 3: 제약조건 플래깅")
    print("=" * 50)
    df_clean = pd.read_excel(WORKING_FILE, sheet_name='_step1_정제데이터', dtype=str)
    df_lab = pd.read_excel(WORKING_FILE, sheet_name='_step2_랩매핑', dtype=str)

    flag_df = step3_flagging.build_flags(df_clean, df_lab)

    n_multi_papers = int(flag_df['is_multi_presenter'].sum())
    n_multi_people = int(
        flag_df[flag_df['is_multi_presenter']]['발표자'].nunique()
    )
    n_special = int(flag_df['is_special_session'].sum())
    n_chair = int(flag_df['is_chair_requested'].sum())

    print(f"  다중 발표자 논문  : {n_multi_papers}건 ({n_multi_people}명)")
    print(f"  특별세션 논문      : {n_special}건")
    print(f"  좌장 신청 논문    : {n_chair}건")

    write_sheet(flag_df, '_step3_플래그')
    print(f"[완료] _step3_플래그 시트 갱신")

    # 사용자 입력 템플릿은 최초 1회만 생성 (이후 사용자 편집물 보존)
    if sheet_exists('입력_좌장Invited'):
        print(f"  입력_좌장Invited 시트 존재 → 보존 (덮어쓰기 안 함)")
    else:
        template = step3_flagging.build_chair_invited_template(flag_df)
        write_sheet(template, '입력_좌장Invited')
        print(f"[완료] 입력_좌장Invited 템플릿 생성 (사용자 편집용)")


def run_step4():
    print("=" * 50)
    print("Step 4: 논문 배정 도우미")
    print("=" * 50)
    df_clean = pd.read_excel(WORKING_FILE, sheet_name='_step1_정제데이터', dtype=str)
    df_lab = pd.read_excel(WORKING_FILE, sheet_name='_step2_랩매핑', dtype=str)

    # 1) 설정_시간대 시트 준비 (세션 컬럼 없으면 구 스키마 → 재생성)
    needs_rebuild_timeslots = True
    if sheet_exists('설정_시간대'):
        df_timeslots = pd.read_excel(WORKING_FILE, sheet_name='설정_시간대', dtype=str)
        if '세션' in df_timeslots.columns:
            needs_rebuild_timeslots = False
            print(f"  설정_시간대 시트 존재 → 사용자 편집본 사용")
        else:
            print(f"  [주의] 설정_시간대 구 스키마 감지(세션 컬럼 없음) → 기본값으로 재생성")
    if needs_rebuild_timeslots:
        df_timeslots = pd.DataFrame(step4_assignment.DEFAULT_TIMESLOTS)
        write_sheet(df_timeslots, '설정_시간대')
        print(f"[완료] 설정_시간대 시트 생성 ({len(df_timeslots)}슬롯, 세션 컬럼 포함)")

    # 2) 입력_논문배정 시트 준비 (세션 컬럼 없으면 재생성)
    needs_rebuild_assign = True
    if sheet_exists('입력_논문배정'):
        df_assign = pd.read_excel(WORKING_FILE, sheet_name='입력_논문배정', dtype=str)
        if '세션' in df_assign.columns and '슬롯길이' in df_assign.columns:
            needs_rebuild_assign = False
            print(f"  입력_논문배정 시트 존재 → 사용자 편집본 사용")
        else:
            print(f"  [주의] 입력_논문배정 구 스키마 감지 → 템플릿 재생성")
    if needs_rebuild_assign:
        df_assign = step4_assignment.build_assignment_template(df_clean, df_lab)
        write_sheet(df_assign, '입력_논문배정')
        print(f"[완료] 입력_논문배정 템플릿 생성 ({len(df_assign)}건)")
    else:
        # 참고 컬럼(발표자/발표형식/발표분야/랩대표)을 정제데이터 기준으로 refresh.
        # 사용자가 채운 배정 컬럼(날짜/세션/발표장/슬롯순서/슬롯길이/비고)은 건드리지 않음.
        lab_lookup = df_lab.set_index('논문번호').to_dict('index') if '논문번호' in df_lab.columns else {}
        clean_lookup = df_clean.set_index('논문번호').to_dict('index') if '논문번호' in df_clean.columns else {}
        refresh_cols = ['발표자', '발표형식', '발표분야', '랩대표']
        refreshed = []
        # (a) 시트2에 없는 pid(취소된 발표) 행 제거
        clean_ids_set = set(df_clean['논문번호'].fillna('').astype(str))
        orphan_mask = ~df_assign['논문번호'].fillna('').astype(str).isin(clean_ids_set | {''})
        orphan_ids = df_assign.loc[orphan_mask, '논문번호'].fillna('').astype(str).tolist()
        if orphan_ids:
            df_assign = df_assign.loc[~orphan_mask].reset_index(drop=True)

        # (b) 시간미배정(날짜/세션/발표장/슬롯순서 중 하나라도 빈 행) 비고='학생경진대회만'
        def _is_empty(v):
            return pd.isna(v) or str(v).strip() in ('', 'nan', 'NaN', 'None')
        unassigned_tag = '학생경진대회만'
        tagged_ids = []
        for idx, row in df_assign.iterrows():
            missing = any(
                _is_empty(row.get(c)) for c in ['날짜', '세션', '발표장', '슬롯순서']
            )
            if missing:
                cur_memo = '' if pd.isna(row.get('비고')) else str(row.get('비고')).strip()
                if cur_memo != unassigned_tag:
                    df_assign.at[idx, '비고'] = unassigned_tag
                    tagged_ids.append(str(row.get('논문번호', '')).strip())

        for idx, row in df_assign.iterrows():
            pid = str(row.get('논문번호', '')).strip()
            if not pid or pid not in clean_lookup:
                continue
            clean_row = clean_lookup[pid]
            lab_info = lab_lookup.get(pid, {})
            changes = {}
            new_vals = {
                '발표자': clean_row.get('발표자', ''),
                '발표형식': clean_row.get('발표형식', ''),
                '발표분야': clean_row.get('발표분야', ''),
                '랩대표': lab_info.get('랩대표', ''),
            }
            for c in refresh_cols:
                if c not in df_assign.columns:
                    continue
                old = '' if pd.isna(row.get(c)) else str(row.get(c))
                new = '' if pd.isna(new_vals[c]) else str(new_vals[c])
                if old != new:
                    df_assign.at[idx, c] = new
                    changes[c] = f"{old} → {new}"
            if changes:
                refreshed.append((pid, changes))

        # 정제데이터에 새로 추가된 논문(시트2 편집으로 재등록된 케이스 등) 머지
        existing_ids = set(df_assign['논문번호'].fillna('').astype(str))
        clean_ids = set(df_clean['논문번호'].fillna('').astype(str))
        new_ids = clean_ids - existing_ids
        if new_ids:
            df_clean_new = df_clean[df_clean['논문번호'].astype(str).isin(new_ids)]
            df_new_rows = step4_assignment.build_assignment_template(df_clean_new, df_lab)
            for col in df_assign.columns:
                if col not in df_new_rows.columns:
                    df_new_rows[col] = ''
            df_new_rows = df_new_rows[df_assign.columns]
            df_assign = pd.concat([df_assign, df_new_rows], ignore_index=True)

        if refreshed or new_ids or orphan_ids or tagged_ids:
            write_sheet(df_assign, '입력_논문배정')
        if orphan_ids:
            print(f"[완료] 입력_논문배정에서 시트2에 없는 고아 행 {len(orphan_ids)}건 삭제: {', '.join(orphan_ids)}")
        if tagged_ids:
            print(f"[완료] 시간미배정 {len(tagged_ids)}건 비고에 '{unassigned_tag}' 기입: {', '.join(tagged_ids[:20])}" + (" ..." if len(tagged_ids) > 20 else ""))
        if refreshed:
            print(f"[완료] 입력_논문배정 참고컬럼 refresh: {len(refreshed)}건")
            for pid, chg in refreshed[:10]:
                print(f"    - {pid}: " + ", ".join(f"{k}({v})" for k, v in chg.items()))
            if len(refreshed) > 10:
                print(f"    ... 외 {len(refreshed) - 10}건")
        if new_ids:
            new_list = ', '.join(sorted(new_ids))
            print(f"[완료] 입력_논문배정에 신규 논문 {len(new_ids)}건 추가: {new_list}")

    # 3) 배정 현황 분석
    status_df, warn_df, abbr_df, stats = step4_assignment.analyze_assignments(
        df_clean, df_lab, df_assign, df_timeslots
    )

    print(f"  총 논문 수           : {stats['total']}")
    print(f"  배정완료             : {stats['assigned']}")
    print(f"  부분배정             : {stats['partial']}")
    print(f"  미배정               : {stats['unassigned']}")
    print(f"  경고                 : {stats['warnings']}건")

    write_sheet(status_df, '_step4_배정현황')
    write_sheet(warn_df, '_step4_충돌경고')
    write_sheet(abbr_df, '_step4_소속축약')
    print(f"[완료] _step4_배정현황, _step4_충돌경고, _step4_소속축약 시트 갱신")

    # 입력_논문배정 드롭다운 (재적용해도 안전)
    apply_assignment_dropdowns(df_timeslots)
    print(f"[완료] 입력_논문배정 드롭다운 적용 (날짜/세션/발표장/슬롯순서/슬롯길이)")


def run_step6():
    print("=" * 50)
    print("Step 6: 최종 포맷팅 및 출력 시트 생성")
    print("=" * 50)
    df_clean = pd.read_excel(WORKING_FILE, sheet_name='_step1_정제데이터', dtype=str)
    df_lab = pd.read_excel(WORKING_FILE, sheet_name='_step2_랩매핑', dtype=str)
    df_assign = pd.read_excel(WORKING_FILE, sheet_name='입력_논문배정', dtype=str)
    df_timeslots = pd.read_excel(WORKING_FILE, sheet_name='설정_시간대', dtype=str)

    df_invited = (
        pd.read_excel(WORKING_FILE, sheet_name='입력_좌장Invited', dtype=str)
        if sheet_exists('입력_좌장Invited') else None
    )

    # 입력_세션좌장 시트 준비 (세션 컬럼 없으면 구 스키마 → 재생성)
    needs_rebuild_sc = True
    if sheet_exists('입력_세션좌장'):
        df_session_chair = pd.read_excel(
            WORKING_FILE, sheet_name='입력_세션좌장', dtype=str
        )
        if '세션' in df_session_chair.columns:
            needs_rebuild_sc = False
            print(f"  입력_세션좌장 시트 존재 → 사용자 편집본 사용")
        else:
            print(f"  [주의] 입력_세션좌장 구 스키마 감지 → 템플릿 재생성")
    if needs_rebuild_sc:
        df_session_chair = step6_formatting.build_session_chair_template(df_timeslots)
        write_sheet(df_session_chair, '입력_세션좌장')
        print(f"[완료] 입력_세션좌장 시트 템플릿 생성 ({len(df_session_chair)}행)")

    # 포스터 분야 블록 연속성 사전 검증
    ctx_for_check = step6_formatting._make_context(
        df_clean, df_lab, df_assign, df_invited, df_timeslots
    )
    contiguity_errors = step6_formatting._validate_poster_field_contiguity(ctx_for_check)
    slot_num_errors = step6_formatting._validate_poster_slot_numbers(ctx_for_check)
    slot_prefix_errors = step6_formatting._validate_poster_slot_prefix(ctx_for_check)
    if contiguity_errors or slot_num_errors or slot_prefix_errors:
        if contiguity_errors:
            print(f"\n[오류] 포스터 분야 블록 연속성 위반 {len(contiguity_errors)}건")
            print("   같은 세션 내에서 분야가 한 번 끝난 뒤 다시 나타납니다.\n")
            for e in contiguity_errors:
                print(
                    f"   - [{e['세션']}] 슬롯 {e['슬롯순서']} "
                    f"{e['발표자']}(pid={e['논문번호']}) "
                    f"분야={e['발표분야']} ← 이전에 {e['이전분야']} 블록이 끼어 있음"
                )
        if slot_num_errors:
            print(f"\n[오류] 포스터 슬롯 숫자 중복 {len(slot_num_errors)}건")
            print("   같은 세션 안에서 같은 번호가 두 번 쓰였습니다. 코드가 중복되므로 차단.\n")
            for e in slot_num_errors:
                print(
                    f"   - [{e['세션']}] 숫자 {e['번호']:02d} 중복: "
                        f"{e['prev_slot']} {e['prev_발표자']}(pid={e['prev_pid']}) <-> "
                    f"{e['cur_slot']} {e['cur_발표자']}(pid={e['cur_pid']})"
                )
        if slot_prefix_errors:
            print(f"\n[오류] 포스터 슬롯 접두어 불일치 {len(slot_prefix_errors)}건")
            print("   슬롯 접두어(A/B/C)가 해당 세션 레터와 다릅니다.\n")
            for e in slot_prefix_errors:
                print(
                    f"   - [{e['세션']}] {e['발표자']}(pid={e['논문번호']}) "
                    f"슬롯={e['슬롯순서']} ← 이 세션은 '{e['expected']}'로 시작해야 함"
                )
        print("\n입력_논문배정 교정 후 재실행하세요.\n")
        return

    # 모든 출력 시트 생성
    outputs, merges, stats = step6_formatting.build_all(
        df_clean, df_lab, df_assign, df_timeslots, df_invited, df_session_chair
    )

    print(f"  구두 발표 명단       : {stats['n_oral_listed']}건")
    print(f"  포스터 발표 명단     : {stats['n_poster_listed']}건")

    # 출력 시트 기록 (매트릭스/세부일정은 헤더 없이)
    headerless_sheets = {
        '5월7일(목)', '5월8일(금)', '발표 세부 일정',
        '포스터 발표 세부 일정', '포스터 세션 배치'
    }
    # 사용자 수동 편집 보존 대상 시트 (자동 생성이 비워두는 셀만 복원)
    preserve_user_cells = {'5월7일(목)', '5월8일(금)'}
    snapshots = {s: snapshot_sheet_cells(s) for s in preserve_user_cells}
    for sheet_name, df_out in outputs.items():
        use_header = sheet_name not in headerless_sheets
        write_sheet(df_out, sheet_name, header=use_header)
        print(f"[완료] '{sheet_name}' 시트 생성/갱신")

    # 이전 시트에만 있던 비-자동 셀 복원 (수작업 추가 보존)
    for sheet_name, snap in snapshots.items():
        restored = restore_empty_cells(sheet_name, snap)
        if restored:
            print(f"[완료] '{sheet_name}' 수동 입력 셀 {restored}개 복원")

    # 매트릭스 셀 병합 적용 (휴식 가로병합, 45분 발표 세로병합 등)
    for sheet_name, sheet_merges in merges.items():
        if sheet_merges:
            apply_merges(sheet_name, sheet_merges)
            print(f"[완료] '{sheet_name}' 셀 병합 {len(sheet_merges)}건 적용")

    # 출력 시트 컬럼 폭 자동 조정
    autofit_columns(sheet_names=list(outputs.keys()))
    if '포스터 세션 배치' in outputs:
        style_poster_session_sheet('포스터 세션 배치')
    print(f"[완료] 출력 시트 컬럼 폭 자동 조정 완료")


# ---------------------------------------------------------------------------

STEPS = {1: run_step1, 2: run_step2, 3: run_step3, 4: run_step4, 6: run_step6}


def main():
    parser = argparse.ArgumentParser(description='프로그램북 생성 스크립트')
    parser.add_argument('--init', action='store_true',
                        help='input_raw.xlsx로부터 작업 파일 초기 생성')
    parser.add_argument('--step', default=None,
                        help='실행할 Step 번호 (1, 2, ... 또는 all)')
    args = parser.parse_args()

    if args.init:
        init_working_file()

    if args.step is None:
        return

    if args.step == 'all':
        for n in sorted(STEPS.keys()):
            STEPS[n]()
    else:
        n = int(args.step)
        if n not in STEPS:
            raise ValueError(f"Step {n}은 아직 구현되지 않았습니다.")
        STEPS[n]()


if __name__ == '__main__':
    main()
